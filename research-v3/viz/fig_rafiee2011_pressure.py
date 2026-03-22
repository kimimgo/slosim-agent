#!/usr/bin/env python3
"""Rafiee & Thiagarajan (2011) — Publication-quality pressure comparison figure.

Calibration (per Kim & Park 2023 Fig.28):
  - Y-shift: match sloshing-period pressure TROUGHS (저점) to experiment
    (sensor-specific, computed from t=2-8s period minima)
  - X-shift: align global max P2 peak to experiment

Layout: 3-row GridSpec
  Row 1: P2 (z=120mm) aligned time series
  Row 2: P1 (z=21mm) aligned time series
  Row 3: Peak bar comparison (time-matched after alignment)
"""
from pathlib import Path
import numpy as np

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec

try:
    import openpyxl
except ImportError:
    raise ImportError("openpyxl required: pip install openpyxl")

# --- Paths ---
PROJECT = Path(__file__).resolve().parents[2]
SUPP = PROJECT / "supplementary"
SIM_BASE = PROJECT / "simulations" / "exp-c" / "rafiee2011"
OUT_DIR = Path(__file__).parent / "plots"
OUT_DIR.mkdir(exist_ok=True)

# --- Colors ---
C_EXP   = '#333333'
C_KIM   = '#F57C00'
C_DP004 = '#1976D2'
C_DP002 = '#388E3C'

T_PERIOD = 2.016


# ============================================================
# Data Loaders
# ============================================================

def load_exp_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    p1_t, p1_p, p2_t, p2_p = [], [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            try:
                p1_t.append(float(row[0]))
                p1_p.append(float(row[1]))
            except (ValueError, TypeError):
                pass
        if len(row) > 10 and row[9] is not None and row[10] is not None:
            try:
                p2_t.append(float(row[9]))
                p2_p.append(float(row[10]))
            except (ValueError, TypeError):
                pass
    wb.close()
    return np.array(p1_t), np.array(p1_p), np.array(p2_t), np.array(p2_p)


def load_kim2023_xlsx(path):
    """Load raw Kim data (no shift applied yet)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    p1_t, p1_p, p2_t, p2_p = [], [], [], []
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if i > 40000:
            break
        if len(row) > 4 and row[3] is not None and row[4] is not None:
            try:
                p1_t.append(float(row[3]))
                p1_p.append(float(row[4]))
            except (ValueError, TypeError):
                pass
        if len(row) > 13 and row[12] is not None and row[13] is not None:
            try:
                p2_t.append(float(row[12]))
                p2_p.append(float(row[13]))
            except (ValueError, TypeError):
                pass
    wb.close()
    return np.array(p1_t), np.array(p1_p), np.array(p2_t), np.array(p2_p)


def load_dsph_csv(csv_path, probe_idx=0):
    times, pressures = [], []
    with open(csv_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith(('#', 'Part', ' ')):
                continue
            parts = line.split(';')
            try:
                times.append(float(parts[1]))
                pressures.append(float(parts[2 + probe_idx]))
            except (ValueError, IndexError):
                continue
    return np.array(times), np.array(pressures)


def downsample(t, p, target_dt=0.01):
    if len(t) < 2:
        return t, p
    dt = t[1] - t[0]
    if dt >= target_dt * 0.9:
        return t, p
    step = max(1, int(target_dt / dt))
    return t[::step], p[::step]


# ============================================================
# Trough & Peak detection
# ============================================================

def find_period_troughs(t, p, period=T_PERIOD, t_start=2.0, t_end=8.0):
    """Find min pressure in each sloshing period (active region)."""
    troughs = []
    t_cur = t_start
    while t_cur + period <= t_end + 0.1:
        mask = (t >= t_cur) & (t < t_cur + period)
        if mask.sum() > 0:
            idx = np.where(mask)[0][np.argmin(p[mask])]
            troughs.append((t[idx], p[idx]))
        t_cur += period
    return troughs


def compute_trough_yshift(exp_t, exp_p, sim_t, sim_p):
    """Compute y-shift to match sloshing troughs: exp_trough_mean - sim_trough_mean."""
    exp_tr = find_period_troughs(exp_t, exp_p)
    sim_tr = find_period_troughs(sim_t, sim_p)
    if not exp_tr or not sim_tr:
        return 0.0, [], []
    exp_mean = np.mean([p for _, p in exp_tr])
    sim_mean = np.mean([p for _, p in sim_tr])
    return exp_mean - sim_mean, exp_tr, sim_tr


def find_global_max_peak(t, p, t_min=1.0):
    mask = t >= t_min
    if mask.sum() == 0:
        return None, None
    idx = np.where(mask)[0][np.argmax(p[mask])]
    return t[idx], p[idx]


def find_peaks_after(t, p, min_height, min_dist_t=1.5, t_min=1.0):
    peaks = []
    for i in range(1, len(p) - 1):
        if t[i] < t_min:
            continue
        if p[i] > min_height and p[i] > p[i-1] and p[i] > p[i+1]:
            if not peaks or (t[i] - peaks[-1][0]) > min_dist_t:
                peaks.append((t[i], p[i], i))
    return peaks


def find_peak_in_window(t, p, t_center, half_window=0.5):
    mask = (t >= t_center - half_window) & (t <= t_center + half_window)
    if mask.sum() == 0:
        return 0.0, t_center
    idx = np.where(mask)[0][np.argmax(p[mask])]
    return p[idx], t[idx]


# ============================================================
# Main
# ============================================================

def main():
    print("=" * 60)
    print("  Rafiee 2011 — Publication Figure")
    print("  (trough-matched y-shift, peak-aligned x-shift)")
    print("=" * 60)

    # --- Load raw data ---
    exp_p1_t, exp_p1_p, exp_p2_t, exp_p2_p = load_exp_xlsx(
        SUPP / "Sloshing_pressure_experiments.xlsx")
    kim_p1_t_raw, kim_p1_p_raw, kim_p2_t_raw, kim_p2_p_raw = load_kim2023_xlsx(
        SUPP / "Sloshing_pressure.xlsx")

    dsph_raw = {}
    for label, dirname, csv_candidates in [
        ('dp004', 'dp004', ['Pressure_full_Press.csv', 'Pressure_Press.csv']),
        ('dp002', 'dp002', ['Pressure_Press.csv']),
    ]:
        for csv_name in csv_candidates:
            csv_path = SIM_BASE / dirname / csv_name
            if csv_path.exists():
                t, p1 = load_dsph_csv(csv_path, probe_idx=0)
                _, p2 = load_dsph_csv(csv_path, probe_idx=1)
                dsph_raw[label] = {
                    'P1_t': t, 'P1_p': p1 / 1000.0,  # Pa → kPa
                    'P2_t': t, 'P2_p': p2 / 1000.0,
                }
                print(f"  DSPH {label}: {len(t)} pts ({csv_name})")
                break

    # --- Compute trough-based Y-shifts (per sensor) ---
    print("\n--- Trough-Based Y-Shift ---")

    # Kim y-shifts
    kim_yshift = {}
    for sensor, et, ep, kt, kp in [
        ('P1', exp_p1_t, exp_p1_p, kim_p1_t_raw, kim_p1_p_raw),
        ('P2', exp_p2_t, exp_p2_p, kim_p2_t_raw, kim_p2_p_raw),
    ]:
        yshift, exp_tr, kim_tr = compute_trough_yshift(et, ep, kt, kp)
        kim_yshift[sensor] = yshift
        exp_mean = np.mean([p for _, p in exp_tr]) if exp_tr else 0
        sim_mean = np.mean([p for _, p in kim_tr]) if kim_tr else 0
        print(f"  Kim {sensor}: exp_trough={exp_mean:.3f}, kim_trough={sim_mean:.3f} "
              f"→ y-shift={yshift:+.3f} kPa")

    # Apply Kim y-shift
    kim_p1_p = kim_p1_p_raw + kim_yshift['P1']
    kim_p2_p = kim_p2_p_raw + kim_yshift['P2']
    kim_p1_t = kim_p1_t_raw
    kim_p2_t = kim_p2_t_raw
    print(f"  Kim after shift: P1=[{kim_p1_p.min():.2f}, {kim_p1_p.max():.2f}], "
          f"P2=[{kim_p2_p.min():.2f}, {kim_p2_p.max():.2f}]")

    # DSPH y-shifts
    dsph = {}
    for label, data in dsph_raw.items():
        dsph[label] = {}
        for sensor, et, ep in [
            ('P1', exp_p1_t, exp_p1_p),
            ('P2', exp_p2_t, exp_p2_p),
        ]:
            yshift, exp_tr, sim_tr = compute_trough_yshift(
                et, ep, data[f'{sensor}_t'], data[f'{sensor}_p'])
            dsph[label][f'{sensor}_t'] = data[f'{sensor}_t']
            dsph[label][f'{sensor}_p'] = data[f'{sensor}_p'] + yshift
            dsph[label][f'{sensor}_yshift'] = yshift
            exp_mean = np.mean([p for _, p in exp_tr]) if exp_tr else 0
            sim_mean = np.mean([p for _, p in sim_tr]) if sim_tr else 0
            print(f"  DSPH {label} {sensor}: trough={sim_mean:.3f} → y-shift={yshift:+.3f} kPa")

    # --- Peak-based X-shift (P2 global max) ---
    print("\n--- Peak Alignment (P2 global max, after y-shift) ---")
    exp_ref_t, exp_ref_p = find_global_max_peak(exp_p2_t, exp_p2_p, t_min=1.0)
    print(f"  Exp P2 max: t={exp_ref_t:.3f}s, P={exp_ref_p:.2f} kPa")

    shifts = {}
    kim_ref_t, kim_ref_p = find_global_max_peak(kim_p2_t, kim_p2_p, t_min=1.0)
    if kim_ref_t is not None:
        shifts['kim'] = exp_ref_t - kim_ref_t
        print(f"  Kim max: t={kim_ref_t:.3f}s, P={kim_ref_p:.2f} → shift={shifts['kim']:+.3f}s")
    else:
        shifts['kim'] = 0.0

    for label in dsph:
        sim_ref_t, sim_ref_p = find_global_max_peak(
            dsph[label]['P2_t'], dsph[label]['P2_p'], t_min=1.0)
        if sim_ref_t is not None:
            shifts[label] = exp_ref_t - sim_ref_t
            print(f"  DSPH {label} max: t={sim_ref_t:.3f}s, P={sim_ref_p:.2f} → "
                  f"shift={shifts[label]:+.3f}s ({abs(shifts[label]/T_PERIOD):.1f}T)")
        else:
            shifts[label] = 0.0

    # --- Peak errors ---
    print("\n--- Time-Matched Peak Errors (trough-calibrated) ---")
    exp_p2_all = find_peaks_after(exp_p2_t, exp_p2_p, min_height=2.0, min_dist_t=1.5)
    exp_p1_all = find_peaks_after(exp_p1_t, exp_p1_p, min_height=2.0, min_dist_t=1.5)

    peak_errors = {}
    all_sources = [
        ('Kim2023', kim_p1_t, kim_p1_p, kim_p2_t, kim_p2_p, shifts.get('kim', 0)),
    ]
    for dl in ['dp004', 'dp002']:
        if dl in dsph:
            all_sources.append((f'DSPH_{dl}',
                                dsph[dl]['P1_t'], dsph[dl]['P1_p'],
                                dsph[dl]['P2_t'], dsph[dl]['P2_p'],
                                shifts.get(dl, 0)))

    for src, sp1t, sp1p, sp2t, sp2p, shift in all_sources:
        for sensor, exp_peaks, st, sp in [
            ('P2', exp_p2_all, sp2t, sp2p),
            ('P1', exp_p1_all, sp1t, sp1p),
        ]:
            errs = []
            for exp_t_pk, exp_p_pk, _ in exp_peaks:
                sim_pk, _ = find_peak_in_window(st + shift, sp, exp_t_pk, 0.5)
                err = abs(sim_pk - exp_p_pk) / exp_p_pk * 100 if exp_p_pk > 0 else 0
                errs.append((exp_p_pk, sim_pk, err))
            peak_errors[f'{src}_{sensor}'] = errs
            if errs:
                avg = np.mean([e[2] for e in errs])
                detail = " | ".join(f"exp={e:.1f} sim={s:.1f} err={r:.0f}%"
                                    for e, s, r in errs)
                print(f"  {src} {sensor}: {detail}  → avg={avg:.1f}%")

    # ========== FIGURE ==========
    fig = plt.figure(figsize=(14, 14))
    gs = gridspec.GridSpec(3, 1, height_ratios=[3, 3, 2], hspace=0.25)

    def get_sources(sensor):
        s = []
        et = exp_p2_t if sensor == 'P2' else exp_p1_t
        ep = exp_p2_p if sensor == 'P2' else exp_p1_p
        s.append(('Experiment', et, ep, C_EXP, 1.3, '-', 0.0, 5))
        kt = kim_p2_t if sensor == 'P2' else kim_p1_t
        kp = kim_p2_p if sensor == 'P2' else kim_p1_p
        s.append(('Kim & Park (2023)', kt, kp, C_KIM, 0.9, '-.', shifts['kim'], 4))
        for dl, color in [('dp004', C_DP004), ('dp002', C_DP002)]:
            if dl in dsph:
                s.append((f'DSPH dp={dl[2:]}mm', dsph[dl][f'{sensor}_t'],
                          dsph[dl][f'{sensor}_p'], color, 0.9, '-', shifts[dl], 3))
        return s

    # Row 1: P2
    ax1 = fig.add_subplot(gs[0])
    for name, t_arr, p_arr, color, lw, ls, shift, zorder in get_sources('P2'):
        t_s = t_arr + shift
        t_ds, p_ds = downsample(t_s, p_arr)
        shift_str = f' ({shift:+.1f}s)' if abs(shift) > 0.01 else ''
        ax1.plot(t_ds, p_ds, ls, color=color, linewidth=lw, alpha=0.8,
                 label=f'{name}{shift_str}', zorder=zorder)

    for t_pk, p_pk, _ in exp_p2_all:
        ax1.plot(t_pk, p_pk, 'k^', markersize=9, zorder=10)
        ax1.annotate(f'{p_pk:.1f}', (t_pk, p_pk),
                     textcoords='offset points', xytext=(10, 5), fontsize=9,
                     fontweight='bold', color=C_EXP)

    if 'DSPH_dp004_P2' in peak_errors:
        dp4s = shifts.get('dp004', 0)
        for (exp_v, sim_v, err), (exp_t_pk, _, _) in zip(
                peak_errors['DSPH_dp004_P2'], exp_p2_all):
            sv, st = find_peak_in_window(
                dsph['dp004']['P2_t'] + dp4s, dsph['dp004']['P2_p'], exp_t_pk, 0.5)
            ax1.plot(st, sv, 'v', color=C_DP004, markersize=8, zorder=10)
            ax1.annotate(f'{sv:.1f}\n({err:.0f}%)', (st, sv),
                         textcoords='offset points', xytext=(-40, -18), fontsize=8,
                         color=C_DP004, fontweight='bold')

    ax1.set_ylabel('Pressure (kPa)', fontsize=12)
    ax1.set_title('P2 — Left Wall Mid-Height (z = 120 mm)', fontsize=13, fontweight='bold')
    ax1.legend(loc='upper left', fontsize=9, framealpha=0.9)
    ax1.grid(True, alpha=0.2)
    ax1.set_xlim(-0.5, 8.5)
    ax1.set_ylim(-2, None)
    for k in range(1, 5):
        ax1.axvline(k * T_PERIOD, color='gray', ls=':', lw=0.5, alpha=0.3)

    # Row 2: P1
    ax2 = fig.add_subplot(gs[1], sharex=ax1)
    for name, t_arr, p_arr, color, lw, ls, shift, zorder in get_sources('P1'):
        t_s = t_arr + shift
        t_ds, p_ds = downsample(t_s, p_arr)
        shift_str = f' ({shift:+.1f}s)' if abs(shift) > 0.01 else ''
        ax2.plot(t_ds, p_ds, ls, color=color, linewidth=lw, alpha=0.8,
                 label=f'{name}{shift_str}', zorder=zorder)

    for t_pk, p_pk, _ in exp_p1_all:
        ax2.plot(t_pk, p_pk, 'k^', markersize=9, zorder=10)
        ax2.annotate(f'{p_pk:.1f}', (t_pk, p_pk),
                     textcoords='offset points', xytext=(10, 5), fontsize=9,
                     fontweight='bold', color=C_EXP)

    ax2.set_ylabel('Pressure (kPa)', fontsize=12)
    ax2.set_xlabel('Time (s)', fontsize=12)
    ax2.set_title('P1 — Left Wall Bottom (z = 21 mm)', fontsize=13, fontweight='bold')
    ax2.legend(loc='upper left', fontsize=9, framealpha=0.9)
    ax2.grid(True, alpha=0.2)
    ax2.set_ylim(-2, None)
    for k in range(1, 5):
        ax2.axvline(k * T_PERIOD, color='gray', ls=':', lw=0.5, alpha=0.3)

    # Row 3: Peak bars
    ax3 = fig.add_subplot(gs[2])
    all_exp_peaks = list(exp_p2_all) + list(exp_p1_all)
    n_p2, n_p1 = len(exp_p2_all), len(exp_p1_all)
    n_total = n_p2 + n_p1
    bar_labels = [f'P2 Peak {i+1}' for i in range(n_p2)] + \
                 [f'P1 Peak {i+1}' for i in range(n_p1)]
    x = np.arange(n_total)
    bw = 0.20

    exp_vals = np.array([p for _, p, _ in all_exp_peaks])
    ax3.bar(x - 1.5*bw, exp_vals, bw, color=C_EXP, alpha=0.65,
            label='Experiment', edgecolor='black', linewidth=0.5)

    kim_vals = []
    for i, (t_pk, _, _) in enumerate(all_exp_peaks):
        sensor = 'P2' if i < n_p2 else 'P1'
        kt = (kim_p2_t if sensor == 'P2' else kim_p1_t) + shifts['kim']
        kp = kim_p2_p if sensor == 'P2' else kim_p1_p
        v, _ = find_peak_in_window(kt, kp, t_pk, 0.5)
        kim_vals.append(v)
    kim_vals = np.array(kim_vals)
    ax3.bar(x - 0.5*bw, kim_vals, bw, color=C_KIM, alpha=0.65,
            label='Kim & Park (2023)', edgecolor='black', linewidth=0.5)

    for dl, color, offset in [('dp004', C_DP004, 0.5), ('dp002', C_DP002, 1.5)]:
        if dl not in dsph:
            continue
        vals = []
        for i, (t_pk, _, _) in enumerate(all_exp_peaks):
            sensor = 'P2' if i < n_p2 else 'P1'
            dt = dsph[dl][f'{sensor}_t'] + shifts[dl]
            dp = dsph[dl][f'{sensor}_p']
            v, _ = find_peak_in_window(dt, dp, t_pk, 0.5)
            vals.append(v)
        vals = np.array(vals)
        ax3.bar(x + offset*bw, vals, bw, color=color, alpha=0.65,
                label=f'DSPH dp={dl[2:]}mm', edgecolor='black', linewidth=0.5)

        if dl == 'dp004':
            for i in range(n_total):
                if exp_vals[i] > 0 and vals[i] > 0:
                    err = abs(vals[i] - exp_vals[i]) / exp_vals[i] * 100
                    y_pos = max(vals[i], exp_vals[i]) + 0.3
                    ax3.annotate(f'{err:.0f}%', (x[i] + offset*bw, y_pos),
                                 ha='center', fontsize=8, fontweight='bold', color=color)

    ax3.set_xticks(x)
    ax3.set_xticklabels(bar_labels, fontsize=9)
    ax3.set_ylabel('Peak Pressure (kPa)', fontsize=11)
    ax3.set_title('Time-Matched Peaks (trough-calibrated, peak-aligned)',
                  fontsize=12, fontweight='bold')
    ax3.legend(loc='upper right', fontsize=8, ncol=2)
    ax3.grid(axis='y', alpha=0.2)
    if n_p2 > 0 and n_p1 > 0:
        ax3.axvline(n_p2 - 0.5, color='gray', ls='--', lw=0.8, alpha=0.4)

    # Summary
    dp4_p2 = peak_errors.get('DSPH_dp004_P2', [])
    dp4_p1 = peak_errors.get('DSPH_dp004_P1', [])
    dp4_p2_avg = np.mean([e[2] for e in dp4_p2]) if dp4_p2 else 0
    dp4_p1_avg = np.mean([e[2] for e in dp4_p1]) if dp4_p1 else 0
    kim_p2_errs = peak_errors.get('Kim2023_P2', [])
    kim_p2_avg = np.mean([e[2] for e in kim_p2_errs]) if kim_p2_errs else 0

    dp4_yP2 = dsph.get('dp004', {}).get('P2_yshift', 0)
    dp4_yP1 = dsph.get('dp004', {}).get('P1_yshift', 0)
    dp4_xs = shifts.get('dp004', 0)

    summary = (
        f"Trough-matched calibration\n"
        f"Kim y-shift: P2={kim_yshift['P2']:+.2f}, P1={kim_yshift['P1']:+.2f}\n"
        f"dp4mm y-shift: P2={dp4_yP2:+.2f}, P1={dp4_yP1:+.2f}\n"
        f"dp4mm x-shift: {dp4_xs:+.2f}s ({abs(dp4_xs/T_PERIOD):.1f}T)\n"
        f"dp4mm P2 err: {dp4_p2_avg:.1f}% | Kim P2: {kim_p2_avg:.1f}%"
    )
    fig.text(0.98, 0.01, summary, fontsize=9, fontfamily='monospace',
             ha='right', va='bottom',
             bbox=dict(boxstyle='round,pad=0.5', facecolor='lightyellow',
                       edgecolor='gray', alpha=0.8))

    fig.suptitle(
        'Rafiee & Thiagarajan (2011) Sloshing Benchmark\n'
        'Trough-Calibrated + Peak-Aligned',
        fontsize=14, fontweight='bold', y=0.99
    )

    out_png = OUT_DIR / "rafiee2011_pressure.png"
    out_pdf = OUT_DIR / "rafiee2011_pressure.pdf"
    fig.savefig(out_png, dpi=200, bbox_inches='tight')
    fig.savefig(out_pdf, bbox_inches='tight')
    plt.close(fig)
    print(f"\n  Saved: {out_png}")
    print(f"         {out_pdf}")
    return str(out_png)


if __name__ == '__main__':
    main()
