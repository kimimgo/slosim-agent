#!/usr/bin/env python3
"""Export ALL SPHERIC figure data to Excel — timeseries only, rho=920 for Oil parametric.

Sheets:
  1. Water_Lateral      — Exp + run_001(dp4) + run_002(dp2) + run_003(dp1)
  2. Oil_Lateral         — Exp + run_005(mDBC) + run_009b(Laminar+SPS) + run_010(probe)
  3. Oil_rho920_a0_5..a3_0 — 26 sheets, one per alpha
  4. Rafiee_P1           — Exp + Kim2023 + DSPH dp4 + dp2
  5. Rafiee_P2           — same
  6. Bridge_Press        — 5-probe pressure from agent-bridge
  7. Baffle_SWL          — baseline vs baffle SWL (Left+Right)
  8. Metadata
"""

from pathlib import Path
import numpy as np

try:
    import openpyxl
except ImportError:
    raise ImportError("pip install openpyxl")

PROJECT = Path(__file__).resolve().parents[2]
SIMDATA = Path("/mnt/simdata/dualsphysics/exp1")
SIMDATA_C = Path("/mnt/simdata/dualsphysics/exp-c")
EXPDATA = PROJECT / "datasets" / "spheric" / "case_1"
SUPP = PROJECT / "supplementary"
SIM_RAFIEE = PROJECT / "simulations" / "exp-c" / "rafiee2011"
SIM_OIL = PROJECT / "simulations"
BRIDGE = PROJECT / "research-v3" / "exp-c" / "agent-bridge"
EXPD = PROJECT / "research-v3" / "exp-d" / "results"
OUT = Path(__file__).parent / "plots" / "SPHERIC_figure_data.xlsx"

T_PERIOD = 2.016


# ---- Loaders ----
def load_exp_ts(filepath):
    data = np.loadtxt(str(filepath), skiprows=1)
    return data[:, 0], data[:, 1] * 100  # t, P(Pa)


def load_dsph_csv(filepath, probe=0):
    times, pressures = [], []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith(("#", "Part", " ")):
                continue
            parts = line.split(";")
            try:
                times.append(float(parts[1]))
                pressures.append(float(parts[2 + probe]))
            except (ValueError, IndexError):
                continue
    return np.array(times), np.array(pressures)


def load_swl_csv(filepath):
    """Load GaugesSWL CSV (semicolon, header row)."""
    times, swlz = [], []
    with open(filepath) as f:
        for i, line in enumerate(f):
            if i == 0:
                continue  # header
            parts = line.strip().split(";")
            try:
                times.append(float(parts[0]))
                swlz.append(float(parts[3]))  # swlz column
            except (ValueError, IndexError):
                continue
    return np.array(times), np.array(swlz)


def smooth(p, w=5):
    if len(p) < w:
        return p
    return np.convolve(p, np.ones(w) / w, mode="same")


def find_peaks_simple(t, p, min_h, min_d):
    from scipy.signal import find_peaks as sp

    ps = smooth(p)
    dt = np.median(np.diff(t)) if len(t) > 1 else 0.001
    dist = max(1, int(min_d / dt))
    pks, _ = sp(ps, height=min_h, distance=dist, prominence=min_h * 0.3)
    valid = [pk for pk in pks if t[pk] > 0.5]
    return [t[pk] for pk in valid], [ps[pk] for pk in valid]


def peak_shift(exp_t, sim_t, n=3):
    n = min(n, len(exp_t), len(sim_t))
    if n == 0:
        return 0.0
    return np.mean([exp_t[i] - sim_t[i] for i in range(n)])


def ds(t, p, hz=1000):
    if len(t) < 2:
        return t, p
    dt = np.median(np.diff(t))
    step = max(1, int(1.0 / (dt * hz)))
    return t[::step], p[::step]


# Rafiee loaders
def load_rafiee_exp(path):
    import openpyxl as xl

    wb = xl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    p1t, p1p, p2t, p2p = [], [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            try:
                p1t.append(float(row[0]))
                p1p.append(float(row[1]))
            except:
                pass
        if len(row) > 10 and row[9] is not None and row[10] is not None:
            try:
                p2t.append(float(row[9]))
                p2p.append(float(row[10]))
            except:
                pass
    wb.close()
    return np.array(p1t), np.array(p1p), np.array(p2t), np.array(p2p)


def load_kim(path):
    import openpyxl as xl

    wb = xl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    p1t, p1p, p2t, p2p = [], [], [], []
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if i > 40000:
            break
        if len(row) > 4 and row[3] is not None and row[4] is not None:
            try:
                p1t.append(float(row[3]))
                p1p.append(float(row[4]))
            except:
                pass
        if len(row) > 13 and row[12] is not None and row[13] is not None:
            try:
                p2t.append(float(row[12]))
                p2p.append(float(row[13]))
            except:
                pass
    wb.close()
    return np.array(p1t), np.array(p1p), np.array(p2t), np.array(p2p)


def trough_yshift(exp_t, exp_p, sim_t, sim_p):
    def troughs(t, p):
        out = []
        tc = 2.0
        while tc + T_PERIOD <= 8.1:
            m = (t >= tc) & (t < tc + T_PERIOD)
            if m.sum() > 0:
                idx = np.where(m)[0][np.argmin(p[m])]
                out.append(p[idx])
            tc += T_PERIOD
        return out

    et, st = troughs(exp_t, exp_p), troughs(sim_t, sim_p)
    if not et or not st:
        return 0.0
    return np.mean(et) - np.mean(st)


def global_max(t, p, tmin=1.0):
    m = t >= tmin
    if m.sum() == 0:
        return None
    return t[np.where(m)[0][np.argmax(p[m])]]


# ---- Write helpers ----
def write_hdr(ws, hdrs):
    for c, h in enumerate(hdrs, 1):
        ws.cell(row=1, column=c, value=h)


def write_col(ws, col, data, row0=2):
    for i, v in enumerate(data):
        ws.cell(row=row0 + i, column=col, value=float(v))


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ============================================================
    # 1. Water Lateral
    # ============================================================
    print("Water Lateral...")
    t_ew, p_ew = load_exp_ts(EXPDATA / "lateral_water_1x.txt")
    ept_w, _ = find_peaks_simple(t_ew, p_ew, 1000, 1.0)

    ws = wb.create_sheet("Water_Lateral")
    hdrs = ["Exp_t(s)", "Exp_P(Pa)"]

    sims_w = [
        ("dp4mm", SIMDATA / "run_001" / "PressLateral_dp004_Press.csv", 0),
        ("dp2mm", SIMDATA / "run_002" / "PressLateral_dp002_Press.csv", 3),
        ("dp1mm", SIMDATA / "run_003" / "PressConsistent_Press.csv", 0),
    ]
    loaded_w = {}
    for label, csv, pi in sims_w:
        if not csv.exists():
            print(f"  SKIP {label}")
            continue
        t, p = load_dsph_csv(csv, probe=pi)
        spt, _ = find_peaks_simple(t, p, 800, 1.0)
        sh = peak_shift(ept_w, spt, 3)
        loaded_w[label] = (t + sh, p)
        hdrs += [f"{label}_t(s)", f"{label}_P(Pa)"]
        print(f"  {label}: {len(t)} pts, shift={sh:+.3f}s")

    write_hdr(ws, hdrs)
    td, pd = ds(t_ew, p_ew)
    write_col(ws, 1, td)
    write_col(ws, 2, pd)
    col = 3
    for label, (t, p) in loaded_w.items():
        td, pd = ds(t, p)
        write_col(ws, col, td)
        write_col(ws, col + 1, pd)
        col += 2

    # ============================================================
    # 2. Oil Lateral
    # ============================================================
    print("Oil Lateral...")
    t_eo, p_eo = load_exp_ts(EXPDATA / "lateral_oil_1x.txt")
    ept_o, _ = find_peaks_simple(t_eo, p_eo, 300, 1.2)

    ws2 = wb.create_sheet("Oil_Lateral")
    hdrs = ["Exp_t(s)", "Exp_P(Pa)"]

    sims_o = [
        ("run005_mDBC", SIMDATA / "run_005" / "PressConsistent_Press.csv", 0),
        (
            "run009b_LamSPS",
            PROJECT
            / "research-v3"
            / "exp-c"
            / "analysis"
            / "run_009b"
            / "pressure_009b.csv",
            0,
        ),
        ("run010_probe", SIMDATA_C / "run_010" / "PressLateral_Press.csv", 0),
    ]
    loaded_o = {}
    for label, csv, pi in sims_o:
        if not csv.exists():
            print(f"  SKIP {label}")
            continue
        t, p = load_dsph_csv(csv, probe=pi)
        spt, _ = find_peaks_simple(t, p, 200, 1.0)
        sh = peak_shift(ept_o[:1], spt[:1], 1)
        loaded_o[label] = (t + sh, p)
        hdrs += [f"{label}_t(s)", f"{label}_P(Pa)"]
        print(f"  {label}: {len(t)} pts, shift={sh:+.3f}s")

    write_hdr(ws2, hdrs)
    td, pd = ds(t_eo, p_eo)
    write_col(ws2, 1, td)
    write_col(ws2, 2, pd)
    col = 3
    for label, (t, p) in loaded_o.items():
        td, pd = ds(t, p)
        write_col(ws2, col, td)
        write_col(ws2, col + 1, pd)
        col += 2

    # ============================================================
    # 3. Oil rho920 parametric (26 alpha values)
    # ============================================================
    print("Oil rho920 parametric...")

    # Load experiment once for all alpha sheets
    exp_oil_file = EXPDATA / "lateral_oil_1x.txt"
    if not exp_oil_file.exists():
        # fallback
        exp_oil_file = PROJECT / "paper-pof" / "data" / "spheric" / "lateral_oil_1x.txt"
    t_exp_oil, p_exp_oil = load_exp_ts(exp_oil_file)
    td_exp, pd_exp = ds(t_exp_oil, p_exp_oil, 500)

    alphas = sorted(
        [
            d.name
            for d in SIM_OIL.iterdir()
            if d.is_dir() and d.name.startswith("oil_rho920_a")
        ]
    )
    print(f"  Found {len(alphas)} alpha cases")

    for adir in alphas:
        # Find CSV
        csv = None
        for candidate in [
            SIM_OIL / adir / "out" / "measure_lateral_Press.csv",
            SIM_OIL / adir / "out" / adir / "measure_lateral_Press.csv",
        ]:
            if candidate.exists():
                csv = candidate
                break
        if csv is None:
            print(f"  SKIP {adir}: no CSV")
            continue

        t, p = load_dsph_csv(csv, probe=0)
        alpha_str = adir.replace("oil_rho920_a", "").replace("_", ".")
        sheet_name = f"rho920_a{alpha_str}"

        ws_a = wb.create_sheet(sheet_name)
        write_hdr(
            ws_a, ["Exp_t(s)", "Exp_P(Pa)", f"a{alpha_str}_t(s)", f"a{alpha_str}_P(Pa)"]
        )
        write_col(ws_a, 1, td_exp)
        write_col(ws_a, 2, pd_exp)
        td, pd = ds(t, p, 500)
        write_col(ws_a, 3, td)
        write_col(ws_a, 4, pd)
        print(f"  {sheet_name}: {len(t)} pts")

    # ============================================================
    # 4-5. Rafiee P1/P2
    # ============================================================
    print("Rafiee 2011...")
    e_p1t, e_p1p, e_p2t, e_p2p = load_rafiee_exp(
        SUPP / "Sloshing_pressure_experiments.xlsx"
    )
    k_p1t, k_p1p_raw, k_p2t, k_p2p_raw = load_kim(SUPP / "Sloshing_pressure.xlsx")
    print(f"  Exp: P1={len(e_p1t)}, P2={len(e_p2t)} pts")
    print(f"  Kim: P1={len(k_p1t)}, P2={len(k_p2t)} pts")

    # Y-shifts
    ky = {}
    for s, et, ep, kt, kp in [
        ("P1", e_p1t, e_p1p, k_p1t, k_p1p_raw),
        ("P2", e_p2t, e_p2p, k_p2t, k_p2p_raw),
    ]:
        ky[s] = trough_yshift(et, ep, kt, kp)
    k_p1p = k_p1p_raw + ky["P1"]
    k_p2p = k_p2p_raw + ky["P2"]

    # DSPH
    dsph = {}
    for label, dname, csvs in [
        ("dp004", "dp004", ["Pressure_full_Press.csv", "Pressure_Press.csv"]),
        ("dp002", "dp002", ["Pressure_Press.csv"]),
    ]:
        for cn in csvs:
            cp = SIM_RAFIEE / dname / cn
            if cp.exists():
                t, p1 = load_dsph_csv(cp, probe=0)
                _, p2 = load_dsph_csv(cp, probe=1)
                dsph[label] = {"t": t, "P1_raw": p1 / 1000, "P2_raw": p2 / 1000}
                print(f"  DSPH {label}: {len(t)} pts")
                break

    for dl in dsph:
        for s, et, ep in [("P1", e_p1t, e_p1p), ("P2", e_p2t, e_p2p)]:
            ys = trough_yshift(et, ep, dsph[dl]["t"], dsph[dl][f"{s}_raw"])
            dsph[dl][s] = dsph[dl][f"{s}_raw"] + ys

    # X-shift
    exp_ref = global_max(e_p2t, e_p2p)
    shifts = {}
    kr = global_max(k_p2t, k_p2p)
    shifts["kim"] = (exp_ref - kr) if kr else 0.0
    for dl in dsph:
        r = global_max(dsph[dl]["t"], dsph[dl]["P2"])
        shifts[dl] = (exp_ref - r) if r else 0.0

    for sensor in ["P2", "P1"]:
        ws_r = wb.create_sheet(f"Rafiee_{sensor}")
        et = e_p2t if sensor == "P2" else e_p1t
        ep = e_p2p if sensor == "P2" else e_p1p
        hdrs = ["Exp_t(s)", f"Exp_{sensor}(kPa)", "Kim_t(s)", f"Kim_{sensor}(kPa)"]
        for dl in ["dp004", "dp002"]:
            if dl in dsph:
                hdrs += [f"{dl}_t(s)", f"{dl}_{sensor}(kPa)"]
        write_hdr(ws_r, hdrs)

        td, pd = ds(et, ep, 500)
        write_col(ws_r, 1, td)
        write_col(ws_r, 2, pd)

        kt = (k_p2t if sensor == "P2" else k_p1t) + shifts["kim"]
        kp = k_p2p if sensor == "P2" else k_p1p
        td, pd = ds(kt, kp, 500)
        write_col(ws_r, 3, td)
        write_col(ws_r, 4, pd)

        col = 5
        for dl in ["dp004", "dp002"]:
            if dl not in dsph:
                continue
            dt = dsph[dl]["t"] + shifts[dl]
            dp = dsph[dl][sensor]
            td, pd = ds(dt, dp, 500)
            write_col(ws_r, col, td)
            write_col(ws_r, col + 1, pd)
            col += 2

    # ============================================================
    # 6. Bridge Press (5 probes)
    # ============================================================
    print("Bridge Press...")
    bp = BRIDGE / "gauges_Press.csv"
    if bp.exists():
        ws_b = wb.create_sheet("Bridge_Press")
        hdrs = ["t(s)"]
        # Parse probe positions from header
        with open(bp) as f:
            lines = f.readlines()
        posx = lines[0].strip().split(";")[2:]  # skip first 2
        n_probes = len(posx)
        for i, x in enumerate(posx):
            hdrs.append(f"Press_{i}_x{x}(Pa)")
        write_hdr(ws_b, hdrs)

        times, probes = [], [[] for _ in range(n_probes)]
        for line in lines[4:]:  # skip 4 header lines
            parts = line.strip().split(";")
            try:
                times.append(float(parts[1]))
                for j in range(n_probes):
                    probes[j].append(float(parts[2 + j]))
            except (ValueError, IndexError):
                continue
        times = np.array(times)
        write_col(ws_b, 1, times)
        for j in range(n_probes):
            write_col(ws_b, j + 2, np.array(probes[j]))
        print(f"  {len(times)} pts, {n_probes} probes")

    # ============================================================
    # 7. Baffle SWL
    # ============================================================
    print("Baffle SWL...")
    ws_bf = wb.create_sheet("Baffle_SWL")
    hdrs = []
    col = 1
    for tag, prefix in [("baseline", "baseline"), ("baffle", "baffle")]:
        for side in ["Left", "Right"]:
            f = EXPD / f"{prefix}_SWL_{side}.csv"
            if not f.exists():
                print(f"  SKIP {f.name}")
                continue
            t, z = load_swl_csv(f)
            lbl = f"{tag}_{side}"
            hdrs += [f"{lbl}_t(s)", f"{lbl}_SWLz(m)"]
            write_col(ws_bf, col, t)
            write_col(ws_bf, col + 1, z)
            col += 2
            print(f"  {lbl}: {len(t)} pts")
    write_hdr(ws_bf, hdrs)

    # ============================================================
    # 8. Metadata
    # ============================================================
    ws_m = wb.create_sheet("Metadata")
    meta = [
        ("File", "SPHERIC_figure_data.xlsx"),
        ("Script", "export_spheric_data.py"),
        ("", ""),
        ("=== SPHERIC Test 10 ===", ""),
        ("Source", "Souto-Iglesias & Botia-Vera (2011)"),
        ("DOI", "10.1016/j.oceaneng.2015.05.013"),
        ("Tank", "900x62x508 mm, quasi-2D"),
        ("Exp trials", "103"),
        ("", ""),
        ("Water Lateral", "h=93mm (18.3%), Pitch T=1.630s, A=4deg, Water rho=998"),
        ("  dp4mm", "run_001, DBC, probe@x=0.013m"),
        ("  dp2mm", "run_002, DBC, probe@x=0.013m"),
        ("  dp1mm", "run_003, DBC, probe@x=0.013m (PressConsistent)"),
        ("", ""),
        ("Oil Lateral", "h=93mm (18.3%), Pitch T=1.535s, A=4deg, Oil rho=990 mu=0.045"),
        ("  run005", "dp=2mm, mDBC, 1kHz, Artificial visc"),
        ("  run009b", "dp=2mm, DBC, 10ms, Laminar+SPS"),
        ("  run010", "dp=2mm, DBC, probe fix @x=0.007m"),
        ("", ""),
        ("Oil rho920 Parametric", "rho=920, dp=0.004, 26 alpha values (0.5~3.0)"),
        ("  Probe", "x=0.007 (dp/2 inward), y=0.031, z=0.093"),
        ("  CSV source", "simulations/oil_rho920_a*/out/measure_lateral_Press.csv"),
        ("  Note", "No time-shift applied (raw simulation time)"),
        ("", ""),
        ("=== Rafiee 2011 ===", ""),
        ("Source", "Rafiee & Thiagarajan (2011)"),
        ("Tank", "1300x100x900 mm, Sway a=0.1m f=0.496Hz"),
        ("P1", "z=21mm (bottom), P2: z=120mm (mid-height)"),
        ("Kim2023", "Kim & Park (2023) PoF — mDBC reference"),
        ("Calibration", "Trough-matched y-shift + P2 peak x-shift"),
        ("Units", "kPa"),
        ("", ""),
        ("=== Bridge ===", ""),
        ("Source", "EXP-C agent-bridge, 5 probes at z=0.05m"),
        ("", ""),
        ("=== Baffle SWL ===", ""),
        ("Source", "EXP-D, MeasureTool SWL gauge"),
        ("Baseline", "0.6x0.3x0.4m, 50% fill, sway f=0.95Hz"),
        ("Baffle", "center vertical, h=0.14m, t=0.016m"),
        ("", ""),
        ("Downsampling", "SPHERIC ~1000Hz, Oil_param/Rafiee ~500Hz"),
        ("Time shift", "Peak-aligned (Water 3-peak, Oil 1-peak)"),
        ("Units", "SPHERIC=Pa, Rafiee=kPa, SWL=m"),
    ]
    for i, (k, v) in enumerate(meta):
        ws_m.cell(row=i + 1, column=1, value=k)
        ws_m.cell(row=i + 1, column=2, value=v)

    wb.save(str(OUT))
    sheets = [ws.title for ws in wb.worksheets]
    print(f"\nSaved: {OUT} ({len(sheets)} sheets)")
    for s in sheets:
        print(f"  - {s}")


if __name__ == "__main__":
    main()
