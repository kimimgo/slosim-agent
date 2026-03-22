#!/usr/bin/env python3
"""Rafiee & Thiagarajan (2011) Sloshing Benchmark — Pressure Comparison.
Compares DualSPHysics DBC/mDBC at dp=0.004/0.002 with:
  - Experimental data (Rafiee 2011)
  - Kim & Park (2023) SPH simulation

Experimental data: supplementary/Sloshing_pressure_experiments.xlsx (P1, P2)
Kim2023 sim data:  supplementary/Sloshing_pressure.xlsx (P1 sim, P2 sim)
DualSPHysics data: MeasureTool pressure CSV
"""
import sys
from pathlib import Path
import numpy as np

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D
except ImportError:
    print("matplotlib required"); sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl"); sys.exit(1)

# --- Paths ---
PROJECT = Path(__file__).parent.parent.parent.parent  # slosim-agent root
SUPP = PROJECT / "supplementary"
SIM_BASE = PROJECT / "simulations" / "exp-c" / "rafiee2011"
OUT_DIR = Path(__file__).parent / "rafiee2011"
OUT_DIR.mkdir(exist_ok=True)

# All 4 simulation variants
SIM_VARIANTS = {
    'dp004_DBC':  SIM_BASE / "dp004",
    'dp004_mDBC': SIM_BASE / "dp004_mdbc",
    'dp002_DBC':  SIM_BASE / "dp002",
    'dp002_mDBC': SIM_BASE / "dp002_mdbc",
}

# Probe indices: 0=P1_left, 1=P2_left, 2=P1_swl, 3=P1_right, 4=P2_right, 5=P2_right_swl
P1_LEFT_IDX = 0
P2_LEFT_IDX = 1


def load_exp_xlsx(path):
    """Load experimental P1 and P2 data from Sloshing_pressure_experiments.xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    p1_t, p1_p, p2_t, p2_p = [], [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            try:
                p1_t.append(float(row[0]))
                p1_p.append(float(row[1]))
            except (ValueError, TypeError):
                pass
        if len(row) > 10 and row[9] is not None and row[10] is not None:
            try:
                p2_t.append(float(row[9]))
                p2_p.append(float(row[10]))
            except (ValueError, TypeError):
                pass
    wb.close()
    return np.array(p1_t), np.array(p1_p), np.array(p2_t), np.array(p2_p)


def load_kim2023_xlsx(path):
    """Load Kim & Park (2023) simulation data from Sloshing_pressure.xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    p1_t, p1_p, p2_t, p2_p = [], [], [], []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i > 40000:
            break
        if len(row) > 4 and row[3] is not None and row[4] is not None:
            try:
                p1_t.append(float(row[3]))
                p1_p.append(float(row[4]))
            except (ValueError, TypeError):
                pass
        if len(row) > 13 and row[12] is not None and row[13] is not None:
            try:
                p2_t.append(float(row[12]))
                p2_p.append(float(row[13]))
            except (ValueError, TypeError):
                pass
    wb.close()
    return np.array(p1_t), np.array(p1_p), np.array(p2_t), np.array(p2_p)


def load_measuretool_csv(csv_path):
    """Load MeasureTool pressure CSV. Returns (times, probe_pressures_array)."""
    times, probes = [], []
    with open(csv_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith(' ') or line.startswith('Part'):
                continue
            parts = line.split(';')
            if len(parts) < 4:
                continue
            try:
                t = float(parts[1])
                vals = [float(p) for p in parts[2:8]]
                times.append(t)
                probes.append(vals)
            except (ValueError, IndexError):
                continue
    return np.array(times), np.array(probes)


def downsample(t, p, target_dt=0.01):
    """Downsample high-frequency data to target_dt interval."""
    if len(t) < 2:
        return t, p
    dt = t[1] - t[0]
    if dt >= target_dt * 0.9:
        return t, p
    step = max(1, int(target_dt / dt))
    return t[::step], p[::step]


def find_all_peaks(t, p, min_height=0, min_distance_t=0.5):
    """Find all local maxima above min_height with minimum distance."""
    peaks = []
    for i in range(1, len(p) - 1):
        if p[i] > min_height and p[i] > p[i-1] and p[i] > p[i+1]:
            if not peaks or (t[i] - t[peaks[-1]]) > min_distance_t:
                peaks.append(i)
    return peaks


def compute_peak_error(exp_t, exp_p, sim_t, sim_p, period=2.016, n_peaks=4):
    """Compute average peak pressure error between experiment and simulation.
    Returns: (avg_error_pct, peak_details_list)
    """
    # Find experimental peaks (per period)
    peaks = []
    for k in range(n_peaks):
        t_start = k * period + 0.5 * period
        t_end = (k + 1) * period + 0.5 * period
        mask = (exp_t >= t_start) & (exp_t <= t_end)
        if mask.sum() == 0:
            continue
        idx = np.argmax(exp_p[mask])
        global_idx = np.where(mask)[0][idx]
        peaks.append((exp_t[global_idx], exp_p[global_idx]))

    if not peaks:
        return None, []

    errors = []
    details = []
    for t_peak, p_exp in peaks:
        window = 0.3
        mask = (sim_t >= t_peak - window) & (sim_t <= t_peak + window)
        if mask.sum() == 0:
            continue
        p_sim_max = np.max(sim_p[mask])
        err_pct = abs(p_sim_max - p_exp) / abs(p_exp) * 100.0
        errors.append(err_pct)
        details.append({
            't_exp': t_peak, 'p_exp': p_exp,
            'p_sim': p_sim_max, 'error_pct': err_pct
        })

    avg_err = np.mean(errors) if errors else None
    return avg_err, details


def compute_peak_magnitude_comparison(exp_t, exp_p, sim_t, sim_p, min_height_kpa=1.0):
    """Phase-independent comparison: find N largest peaks in each signal
    and compare sorted magnitudes."""
    exp_peaks = find_all_peaks(exp_t, exp_p, min_height=min_height_kpa, min_distance_t=1.0)
    sim_peaks = find_all_peaks(sim_t, sim_p, min_height=min_height_kpa, min_distance_t=1.0)

    exp_peak_vals = sorted([exp_p[i] for i in exp_peaks], reverse=True)
    sim_peak_vals = sorted([sim_p[i] for i in sim_peaks], reverse=True)

    return exp_peak_vals, sim_peak_vals, exp_peaks, sim_peaks


def main():
    print("=" * 70)
    print("  Rafiee & Thiagarajan (2011) Sloshing Benchmark — 4-Way Comparison")
    print("=" * 70)

    # 1. Load experimental data
    exp_file = SUPP / "Sloshing_pressure_experiments.xlsx"
    if not exp_file.exists():
        print(f"  ERROR: {exp_file} not found"); sys.exit(1)
    print(f"\n  Loading experimental data...")
    exp_p1_t, exp_p1_p, exp_p2_t, exp_p2_p = load_exp_xlsx(exp_file)
    print(f"    P1: {len(exp_p1_t)} pts, P2: {len(exp_p2_t)} pts")

    # 2. Load Kim & Park (2023) simulation
    kim_file = SUPP / "Sloshing_pressure.xlsx"
    has_kim = False
    if kim_file.exists():
        print(f"  Loading Kim & Park (2023) simulation...")
        kim_p1_t, kim_p1_p, kim_p2_t, kim_p2_p = load_kim2023_xlsx(kim_file)
        print(f"    P1 sim: {len(kim_p1_t)} pts, P2 sim: {len(kim_p2_t)} pts")
        has_kim = True

    # 3. Load DualSPHysics data (all available variants)
    dsph_data = {}
    for label, sim_dir in SIM_VARIANTS.items():
        # Try multiple CSV name patterns
        for csv_name in ['Pressure_full_Press.csv', 'Pressure_Press.csv']:
            csv_path = sim_dir / csv_name
            if csv_path.exists():
                t, probes = load_measuretool_csv(csv_path)
                if len(t) > 100:
                    dsph_data[label] = {
                        'time': t, 'probes': probes,
                        'P1': probes[:, P1_LEFT_IDX],
                        'P2': probes[:, P2_LEFT_IDX],
                    }
                    print(f"  DualSPHysics {label}: {len(t)} pts ({csv_name})")
                    break

    if not dsph_data:
        print("\n  WARNING: No DualSPHysics data found. Plotting exp + Kim2023 only.")

    # 4. Pressure time series plots (2 rows: P1, P2)
    fig, axes = plt.subplots(2, 1, figsize=(14, 10), sharex=True)
    colors = {
        'dp004_DBC':  '#2ecc71',
        'dp004_mDBC': '#e74c3c',
        'dp002_DBC':  '#3498db',
        'dp002_mDBC': '#9b59b6',
    }
    styles = {
        'dp004_DBC':  '-',
        'dp004_mDBC': '--',
        'dp002_DBC':  '-',
        'dp002_mDBC': '--',
    }

    for idx, (sensor, exp_t, exp_p, probe_idx) in enumerate([
        ('P1', exp_p1_t, exp_p1_p, P1_LEFT_IDX),
        ('P2', exp_p2_t, exp_p2_p, P2_LEFT_IDX),
    ]):
        ax = axes[idx]

        # Experimental (black)
        ax.plot(exp_t, exp_p, 'k-', linewidth=1.5, alpha=0.8, label='Experiment')

        # Kim & Park 2023 (blue dashed)
        if has_kim:
            kim_t = kim_p1_t if sensor == 'P1' else kim_p2_t
            kim_p = kim_p1_p if sensor == 'P1' else kim_p2_p
            ax.plot(kim_t, kim_p, color='#f39c12', linewidth=1.0, alpha=0.7,
                    linestyle='-.', label='Kim & Park (2023)')

        # DualSPHysics variants
        for label, data in dsph_data.items():
            t_ds, p_ds = downsample(data['time'], data['probes'][:, probe_idx])
            p_kpa = p_ds / 1000.0  # Pa to kPa
            ax.plot(t_ds, p_kpa, color=colors[label], linestyle=styles[label],
                    linewidth=0.8, alpha=0.7, label=f'DSPH {label}')

        ax.set_ylabel('Pressure (kPa)', fontsize=12)
        ax.set_title(f'{sensor} — Wall Pressure', fontsize=13)
        ax.legend(loc='upper right', fontsize=8, ncol=2)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(-0.2, 8.2)

    axes[1].set_xlabel('Time (s)', fontsize=12)
    fig.suptitle('Rafiee & Thiagarajan (2011) Sloshing — Resolution Comparison\n'
                 'Experiment vs Kim & Park (2023) vs DualSPHysics DBC (dp=0.004, 0.002)',
                 fontsize=13, fontweight='bold')
    fig.tight_layout()

    out_png = OUT_DIR / 'rafiee2011_pressure_4way.png'
    out_pdf = OUT_DIR / 'rafiee2011_pressure_4way.pdf'
    fig.savefig(out_png, dpi=150)
    fig.savefig(out_pdf, bbox_inches='tight')
    print(f"\n  Saved: {out_png}")
    print(f"         {out_pdf}")
    plt.close()

    # 5. Quantitative error analysis
    print("\n" + "=" * 70)
    print("  Peak Pressure Error Analysis (vs Experiment)")
    print("=" * 70)

    results = {}
    for sensor, exp_t, exp_p, probe_idx in [
        ('P1', exp_p1_t, exp_p1_p, P1_LEFT_IDX),
        ('P2', exp_p2_t, exp_p2_p, P2_LEFT_IDX),
    ]:
        print(f"\n  {sensor}:")

        # Kim & Park (2023)
        if has_kim:
            kim_t = kim_p1_t if sensor == 'P1' else kim_p2_t
            kim_p_arr = (kim_p1_p if sensor == 'P1' else kim_p2_p)
            avg_err, details = compute_peak_error(exp_t, exp_p, kim_t, kim_p_arr)
            if avg_err is not None:
                print(f"    Kim & Park (2023):  avg peak error = {avg_err:.1f}%")
                results[f'Kim2023_{sensor}'] = avg_err

        # DualSPHysics variants
        for label, data in dsph_data.items():
            sim_p_kpa = data['probes'][:, probe_idx] / 1000.0
            avg_err, details = compute_peak_error(exp_t, exp_p, data['time'], sim_p_kpa)
            if avg_err is not None:
                print(f"    DSPH {label:15s}: avg peak error = {avg_err:.1f}%")
                for d in details:
                    print(f"      t={d['t_exp']:.2f}s: exp={d['p_exp']:.2f} sim={d['p_sim']:.2f} kPa  err={d['error_pct']:.1f}%")
                results[f'{label}_{sensor}'] = avg_err

    # 6. Summary table
    print("\n" + "=" * 70)
    print("  Summary Error Table (avg peak error %)")
    print("=" * 70)
    print(f"  {'Method':20s} {'P1':>10s} {'P2':>10s} {'Mean':>10s}")
    print(f"  {'-'*20} {'-'*10} {'-'*10} {'-'*10}")

    methods = ['Kim2023']
    for v in SIM_VARIANTS:
        methods.append(v)

    for m in methods:
        p1_err = results.get(f'{m}_P1', None)
        p2_err = results.get(f'{m}_P2', None)
        p1_str = f"{p1_err:.1f}%" if p1_err is not None else "N/A"
        p2_str = f"{p2_err:.1f}%" if p2_err is not None else "N/A"
        if p1_err is not None and p2_err is not None:
            mean_str = f"{(p1_err + p2_err)/2:.1f}%"
        else:
            mean_str = "N/A"
        print(f"  {m:20s} {p1_str:>10s} {p2_str:>10s} {mean_str:>10s}")

    # 7. Phase-independent peak magnitude comparison
    print("\n" + "=" * 70)
    print("  Phase-Independent Peak Comparison (sorted magnitudes)")
    print("=" * 70)

    for sensor, exp_t, exp_p, probe_idx in [
        ('P1', exp_p1_t, exp_p1_p, P1_LEFT_IDX),
        ('P2', exp_p2_t, exp_p2_p, P2_LEFT_IDX),
    ]:
        print(f"\n  {sensor}:")
        exp_peaks = find_all_peaks(exp_t, exp_p, min_height=1.0, min_distance_t=1.0)
        print(f"    Experiment peaks: {[f'{exp_p[i]:.2f}kPa@t={exp_t[i]:.2f}' for i in exp_peaks]}")

        for label, data in dsph_data.items():
            sim_p_kpa = data['probes'][:, probe_idx] / 1000.0
            sim_peaks = find_all_peaks(data['time'], sim_p_kpa, min_height=0.5, min_distance_t=1.0)
            if sim_peaks:
                sim_t_arr = data['time']
                peak_strs = [f'{sim_p_kpa[i]:.2f}kPa@t={sim_t_arr[i]:.2f}' for i in sim_peaks]
                print(f"    {label:15s}: {peak_strs}")
                # Max peak comparison
                exp_max = max(exp_p[i] for i in exp_peaks) if exp_peaks else 0
                sim_max = max(sim_p_kpa[i] for i in sim_peaks)
                err = abs(sim_max - exp_max) / exp_max * 100 if exp_max > 0 else 0
                print(f"      Max peak: exp={exp_max:.2f} sim={sim_max:.2f} kPa → err={err:.1f}%")
            else:
                print(f"    {label:15s}: No peaks > 0.5 kPa found")

    # 8. Statistics summary
    print("\n" + "=" * 70)
    print("  Pressure Statistics (t=[2,8]s window)")
    print("=" * 70)

    for sensor, probe_idx in [('P1', P1_LEFT_IDX), ('P2', P2_LEFT_IDX)]:
        print(f"\n  {sensor}:")
        for label, data in dsph_data.items():
            t = data['time']
            p = data['probes'][:, probe_idx] / 1000.0
            mask = (t >= 2.0) & (t <= 8.0)
            if mask.sum() > 0:
                p_window = p[mask]
                p_nonzero = p_window[p_window > 0]
                print(f"    {label:15s}: max={p_window.max():.2f} kPa, "
                      f"mean(nonzero)={p_nonzero.mean():.2f} kPa ({len(p_nonzero)}/{mask.sum()} samples), "
                      f"global_max={p.max():.2f} kPa")

    print(f"\n  [Analysis complete]")


if __name__ == '__main__':
    main()
